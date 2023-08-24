import os
import sys

import openpyxl


from PyQt5 import QtCore, QtGui, QtWidgets

from PyQt5.QtWidgets import QMessageBox, QDialog

from src.authentication import frontend_auth

from src.qt_creating import main_ui
from src.qt_creating import start_ui

class AuthWindow(QtWidgets.QMainWindow, frontend_auth.Ui_WelcomeWindow):

    def __init__(self):
        '''ИНИЦИАЛИЗАЦИЯ'''

        super().__init__()

        self.setupUi(self)

        self.path_to_bd_xlsx = '\\'.join(os.getcwd().split('\\')[0:-1]) + '\\FinalProject\\src\\authentication\\database.xlsx'

        self.write_username()

        '''PUSH BUTTONS ACTION'''
        self.junction_boxes_pushButton.clicked.connect(self.show_jb_qt)


    def create_dict_for_verification(self):
        '''
        Создает словарь для верификации, определяет имя и почту
        :return:
        '''
        wb = openpyxl.load_workbook(self.path_to_bd_xlsx)
        ws = wb.active
        self.dict_first_second_name = dict()
        for cell in ws['F']:
            if cell.value:
                if '@' in cell.value:
                    value = cell.value.split('@')[0]
                    full_name = ws[f'B{cell.row}'].value
                    first_name = full_name.split(' ')[0]
                    second_name = full_name.split(' ')[1]
                    third_name = None
                    if len(full_name.split(' ')) >= 3:
                        third_name = full_name.split(' ')[2]
                    if third_name:
                        self.dict_first_second_name[value] = first_name + f' {second_name[0]}.{third_name[0]}.'
                    else:
                        self.dict_first_second_name[value] = first_name + f' {second_name[0]}.'

    def write_username(self):
        '''
        Записываем кто разработал
        :return:
        '''
        self.create_dict_for_verification()

        if (os.getlogin() != '' and os.getlogin() != 'admin'):
            if os.getlogin() in list(self.dict_first_second_name.keys()):
                self.username_QLineEdit.insert(self.dict_first_second_name[os.getlogin()])

    def return_username(self):
        '''Возвращает username'''
        return self.username_QLineEdit.text()

    def return_task_number(self):
        '''Возвращает номер заявки'''
        return self.number_task_QLineEdit.text()

    def return_position_number(self):
        '''Возвращает номер заявки'''
        return self.position_number_QLineEdit.text()

    def set_task_number(self,task_number):
        '''Устанавливает номер задачи'''
        self.number_task_QLineEdit.setText(task_number)

    def set_position_number(self,position_number):
        '''Устанавливает номер задачи'''
        self.position_number_QLineEdit.setText(position_number)


    def show_jb_qt(self):
        path_to_csv = '\\'.join(os.getcwd().split('\\')[0:-1]) + '\\bd'
        path_to_dxf_shell = '\\'.join(os.getcwd().split('\\')[0:-1]) + '\\dxf_base\\DXF_BASE.dxf'
        path_to_terminal_dxf = '\\'.join(os.getcwd().split('\\')[0:-1]) + '\\dxf_base\\DXF_BASE.dxf'
        self.jb_window = main_ui.DxfCreator(path_to_csv=path_to_csv,
                                            path_to_dxf=path_to_dxf_shell,
                                            path_to_terminal_dxf=path_to_terminal_dxf)
        # self.jb_window = start_ui.Mainver(path_to_terminal_dxf=path_to_terminal_dxf)
        self.jb_window.designer_name = self.return_username()
        self.jb_window.task_number = self.return_task_number()
        self.jb_window.position_number = self.return_position_number()
        self.close()
        self.jb_window.show()

def run_app():
    app = QtWidgets.QApplication(sys.argv)
    authWindow = AuthWindow()
    authWindow.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    authWindow = AuthWindow()
    authWindow.show()
    sys.exit(app.exec_())
