import math

import ezdxf
import openpyxl
from openpyxl.utils import get_column_letter


def create_doc_BOM(dxfbase_path:str):
    '''Создает BOM удаляя все не нужное'''
    doc_bom = ezdxf.readfile(dxfbase_path)

    doc_dxfbase_for_del = ezdxf.readfile(dxfbase_path)

    doc_bom.modelspace().delete_all_entities()

    blocks_BOM = ['BOM_FIRST','BOM_SECOND']


    for block in doc_dxfbase_for_del.blocks:
        try:
            if block.dxf.name not in blocks_BOM and '*' not in block.dxf.name:
                    doc_bom.blocks.delete_block(name=block.dxf.name)
        except:
            continue
    return doc_bom

def create_BOM_FIRST(doc_bom):
    '''Создает первый лист спецификации'''

    block_border = doc_bom.blocks['BOM_FIRST']
    values = {attdef.dxf.tag: '' for attdef in block_border.query('ATTDEF')}
    if doc_bom.blocks.get('BOM_FIRST'):
        border_insert = doc_bom.modelspace().add_blockref(name='BOM_FIRST',
                                                          insert=(0, 0))
        border_insert.add_auto_attribs(values)

        return border_insert

def check_name_len(name_string:str):
    if name_string != None:
        if math.ceil(len(name_string)/27) ==1:
            return True
        else:
            return False

def read_BOM_base(xlsx_base_path:str):
    '''
    Чтение базавой эксельки BOM и выдача словаря
    :param xlsx_base_path:
    :return:
    '''

    return_dict = dict()

    workbook = openpyxl.load_workbook(xlsx_base_path)
    worksheet = workbook.active

    column_names = list(worksheet[1])

    for column_name in column_names:
        return_dict[column_name.value] = {}
        for count,row_value in enumerate(worksheet[get_column_letter(column_name.column)][column_name.row:]):
            return_dict[column_name.value][count] = row_value.value

    return return_dict

def write_attrib(BOM_insert,dict_for_writing_attrib):
    '''
    Заполение аттрибутов в спецификации
    :param BOM_insert: BOM_FIRST или BOM_SECOND
    :param dict_for_writing_attrib:
    {'Сборочные единицы':{0:{"Формат":"А4","Зона":"","Поз.":0+1,"Обозначение":"ВРПТ.301172.024-11","Наименование":"Оболочка ВП.161610", "Кол.":1, "Примечание": производитель}}
    :return:
    '''

    row_start = 2
    dict_name_attrib = {attrib.dxf.tag: attrib for attrib in BOM_insert.attribs}
    for main_type_name in dict_for_writing_attrib:
        # [Сборочные изделия, Детали, Стандартные изделия ...]
        dict_name_attrib[f'E{row_start}'].dxf.text = main_type_name
        row_start+=2

        dict_with_pozition = dict_for_writing_attrib[main_type_name]

        for number_pozition in dict_with_pozition:
            dict_name_attrib[f'A{row_start}'].dxf.text = dict_with_pozition[number_pozition]['Формат']
            dict_name_attrib[f'B{row_start}'].dxf.text = dict_with_pozition[number_pozition]['Зона']
            dict_name_attrib[f'C{row_start}'].dxf.text = dict_with_pozition[number_pozition]['Поз.']
            dict_name_attrib[f'D{row_start}'].dxf.text = dict_with_pozition[number_pozition]['Обозначение']
            dict_name_attrib[f'F{row_start}'].dxf.text = dict_with_pozition[number_pozition]['Кол.']
            dict_name_attrib[f'G{row_start}'].dxf.text = dict_with_pozition[number_pozition]['Примечание']

            if math.ceil(len(dict_with_pozition[number_pozition]['Наименование']) / 27) == 1:
                dict_name_attrib[f'E{row_start}'].dxf.text = dict_with_pozition[number_pozition]['Наименование']
            elif math.ceil(len(dict_with_pozition[number_pozition]['Наименование']) / 27) == 2:
                dict_name_attrib[f'E{row_start}'].dxf.text = dict_with_pozition[number_pozition]['Наименование'][:27]
                row_start+=1
                dict_name_attrib[f'E{row_start}'].dxf.text = dict_with_pozition[number_pozition]['Наименование'][27:]

    row_start+=1

def calculate_count_row_for_one_position(dict_with_all_info_in_BOM_row:dict)->int:
    '''
    Расчитать количество строк, которые понадобятся для этой позиции для заполнения в BOM
    :param dict_with_all_info_in_BOM_row: {'Обозначение': 'ВРПТ.301172.024-11', 'Наименование': 'Кабельный ввод ВЗ-Н25#для не бронированного#кабеля, диаметром 12-18мм',
                                            'Формат': 'А4', 'Кол.': None, 'Примечание': None}
    :return: row:int
    '''
    row = 0
    for column_name, name_in_bom in dict_with_all_info_in_BOM_row.items():
        if name_in_bom== None:
            name_in_bom = ''
        name_in_bom = str(name_in_bom)
        if '#' not in name_in_bom:
            row = max(1, row)
        else:
            row = max(row, len(name_in_bom.split('#')))
    return row

def recreate_dict_with_all_info_in_BOM_row(dict_with_all_info_in_BOM_row:dict,row:int):
    '''
    Пересоздает словарь dict_with_all_info_in_BOM_row
    :param dict_with_all_info_in_BOM_row: {'Обозначение': 'ВРПТ.301172.024-11', 'Наименование': 'Кабельный ввод ВЗ-Н25#для не бронированного#кабеля, диаметром 12-18мм',
                                            'Формат': 'А4', 'Кол.': None, 'Примечание': None}
    :param row: 4
    :return: {'Обозначение': ['ВРПТ.301172.024-11','',''], 'Наименование': ['Кабельный ввод ВЗ-Н25','для не бронированного','кабеля, диаметром 12-18мм'],
                'Формат': ['А4','',''], 'Кол.': ['','',''], 'Примечание': ['','','']}
    '''
    for column_name, name_in_bom in dict_with_all_info_in_BOM_row.items():
        if name_in_bom== None:
            name_in_bom = ''
        if '#' not in name_in_bom:
             for i in range(0, row):
                 if i == 0:
                     dict_with_all_info_in_BOM_row[column_name] = [name_in_bom]
                 else:
                     dict_with_all_info_in_BOM_row[column_name].append('')
        else:
            row_new = name_in_bom.split('#')
            for i in range(0, row):
                if len(row_new) == row:
                    dict_with_all_info_in_BOM_row[column_name] = row_new
                else:
                    dict_with_all_info_in_BOM_row[column_name] = row_new
                    for i in range(0,row - len(row_new)):
                        dict_with_all_info_in_BOM_row[column_name].append('')


def add_dict(dict_with_all_info_in_BOM_row:dict, count_row:int):
    '''
    Нужно добавить в этот же словарь тэг того, сколько страниц добавить в конце после прохода по данному словорю а также найти, сколько нужно добавить
    :param dict_with_all_info_in_BOM_row:{'Обозначение': 'ВРПТ.305311.001-025', 'Наименование': 'Кабельный ввод ВЗ-Н25#для не бронированного#кабеля, диаметром 12-18мм', 'Свойство': 'Сборочные единицы', 'Формат': 'А4', 'Кол.': None, 'Примечание': None}
    :param count_row:1
    :return:
    '''
    #Получаем необходимо число строк
    row = calculate_count_row_for_one_position(dict_with_all_info_in_BOM_row)

    for column_name, name_in_bom in dict_with_all_info_in_BOM_row.items():
        if name_in_bom== None:
            name_in_bom = ''
        name_in_bom = str(name_in_bom)
        if '#' not in name_in_bom:
             for i in range(0, row):
                 if i == 0:
                     dict_with_all_info_in_BOM_row[column_name] = [name_in_bom]
                 else:
                     dict_with_all_info_in_BOM_row[column_name].append('')
        else:
            row_new = name_in_bom.split('#')
            for i in range(0, row):
                if len(row_new) == row:
                    dict_with_all_info_in_BOM_row[column_name] = row_new
                else:
                    dict_with_all_info_in_BOM_row[column_name] = row_new
                    for i in range(0,row - len(row_new)):
                        dict_with_all_info_in_BOM_row[column_name].append('')

    return row + count_row

def create_dict_main_properties(list_properties:list):
    '''
    Свойства нужно вынести наружу и сделать словарь
    :param list_properties:
[{'Обозначение': 'ВРПТ.301172.024-11', 'Наименование': 'Оболочка ВП.161610', 'Свойство': 'Сборочные единицы', 'Формат': 'А4', 'Кол.': None, 'Примечание': None},
 {'Обозначение': None, 'Наименование': 'Винт А2.М6-6gx10.019#ГОСТ 17473-80', 'Свойство': 'Стандартные изделия', 'Формат': 'А4', 'Кол.': None, 'Примечание': None},
 {'Обозначение': None, 'Наименование': 'Шайба 6 019 ГОСТ 6402-70', 'Свойство': 'Стандартные изделия', 'Формат': 'А4', 'Кол.': None, 'Примечание': None},
 {'Обозначение': None, 'Наименование': 'Шайба A.6.019 ГОСТ 11371-78', 'Свойство': 'Стандартные изделия', 'Формат': 'А4', 'Кол.': None, 'Примечание': None},
 {'Обозначение': 'ВРПТ.745551.005-140', 'Наименование': 'DIN-рейка NS35х7,5, L=140 мм', 'Свойство': 'Детали', 'Формат': 'А4', 'Кол.': None, 'Примечание': None},
 {'Обозначение': 'ВРПТ.305311.001-025', 'Наименование': 'Кабельный ввод ВЗ-Н25#для не бронированного#кабеля, диаметром 12-18мм', 'Свойство': 'Сборочные единицы', 'Формат': 'А4', 'Кол.': None, 'Примечание': None}]

    :return: "Стандартные изделия":{}
    '''
    return_dict = dict()
    for equip_dict in list_properties:
        property = equip_dict['Свойство']
        if property not in return_dict:
            equip_dict.pop('Свойство')
            return_dict[property] = [equip_dict]
        else:
            equip_dict.pop('Свойство')
            return_dict[property].append(equip_dict)
    return return_dict

def create_list_all_block_names_in_doc(doc_new):
    '''
    Создает список всех блоков импортированных в self.doc_new
    :param doc: self.doc_new
    :return: ['SUPU_SCREW..., VP.161610_topside,...']
    '''
    return  [block.dxf.name for block in doc.blocks if '*' not in block.dxf.name]


def check_next_page(BOM_insert_name:str, row_number:int):
    '''
    Проверка на создание следующей страницы
    :param BOM_insert_name: либо BOM_FIRST либо BOM_SECOND
    :param row_number: 1-29 или 1-32
    :return: True or False
    '''

    if BOM_insert_name == 'BOM_FIRST':
        if row_number > 29:
            return False
        else:
            return True
    elif BOM_insert_name == 'BOM_SECOND':
        if row_number > 32:
            return False
        else:
            return True

def write_mainproperty_in_bom_E_cell(BOM_insert_name:str, row_number:int, mainproperty_attribtag_name:str, dict_attribs:dict):
    '''
    Заполнение спецификации Сборочная единица, Деталь, и тд
    :return:True or False
    '''
    if check_next_page(BOM_insert_name=BOM_insert_name, row_number = row_number + 4):
        row_number +=1
        tag_attrib = 'E' + str(row_number)
        dict_attribs[tag_attrib].dxf.text = mainproperty_attribtag_name
        row_number += 2
        return True
    else:
        return False






if __name__ == '__main__':

    # create_BOM_FIRST(doc_bom=doc)
    # doc.saveas('C:\\Users\\g.zubkov\\PycharmProjects\\FinalProject\\src\\dxf_base\\test_BOM.dxf')

    data_base_bom = read_BOM_base(xlsx_base_path='C:\\Users\\g.zubkov\\PycharmProjects\\FinalProject\\src\\dxf_base\\naming_base.xlsx')
    doc = ezdxf.readfile('C:\\Users\\g.zubkov\\PycharmProjects\\FinalProject\\src\\xx.dxf')

    tag_in_BOM_dxf = {'Формат':'A', 'Зона':'B', 'Поз.':'C', 'Обозначение':'D', 'Наименование':'E', 'Кол.': 'F', 'Примечание':'G'}

    list_with_block_names = create_list_all_block_names_in_doc(doc_new=doc)

    list_for_creating_BOM = list()
    for block_name in list_with_block_names:
        for count_block, name_block_base in data_base_bom['Блок'].items():
            if name_block_base == block_name:
                _dict = {}
                for _ in data_base_bom:
                    if _ != 'Блок':
                        _dict[_]= data_base_bom[_][count_block]
                list_for_creating_BOM.append(_dict)

    doc_bom = create_doc_BOM(dxfbase_path='C:\\Users\\g.zubkov\\PycharmProjects\\FinalProject\src\dxf_base\\DXF_BASE.dxf')

    main_border = create_BOM_FIRST(doc_bom=doc_bom)

    dict_attribs = {attrib.dxf.tag: attrib for attrib in main_border.attribs}

    dict_for_creating_BOM_with = create_dict_main_properties(list_for_creating_BOM)

    start_row_int = 1
    startstart_row_int = 1
    for name_property in dict_for_creating_BOM_with:
        # if write_mainproperty_in_bom_E_cell(BOM_insert_name = main_border, row_number:int, mainproperty_attribtag_name:str, dict_attribs:dict)
        start_row_int += 1
        startstart_row_int += 1
        tag_attrib = 'E' + str(start_row_int)
        dict_attribs[tag_attrib].dxf.text = name_property
        start_row_int +=2
        startstart_row_int +=2
        list_for_creating_BOM = dict_for_creating_BOM_with[name_property]
        for equip_dict in list_for_creating_BOM:
            max_row = add_dict(dict_with_all_info_in_BOM_row=equip_dict, count_row=startstart_row_int)

            for column_name in equip_dict:
                for name in equip_dict[column_name]:
                    if column_name in tag_in_BOM_dxf:
                        tag_attrib = tag_in_BOM_dxf[column_name] + str(start_row_int)
                        if tag_attrib in dict_attribs:
                            dict_attribs[tag_attrib].dxf.text = name
                            start_row_int +=1
                start_row_int = startstart_row_int
            start_row_int = max_row
            startstart_row_int = max_row

    doc_bom.saveas('C:\\Users\\g.zubkov\\PycharmProjects\\FinalProject\src\\bom_check.dxf')

        # for namename in equip:
        #     end_row_int = add_dict(equip,start_row_int)

